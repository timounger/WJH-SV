# This Python file uses the following encoding: utf-8
"""
*****************************************************************************
 @file    wjh_sv.py
 @brief   WJH-SV main file
*****************************************************************************
"""

import sys
import os
import argparse
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

B_DEBUG = True

# Example and Debug Parameter
I_YEAR = 2022
S_FILE = "Tabelle_2022.xlsx"
S_SHEET = "Tabelle"

S_OUTPUT_FOLDER = "Output"

I_EMPTY_ROW_INX = 0
I_HEADER_ROW_INX = 1

# Column position of parameter in input file
ZBNachname = 0
ZBVorname = 1
ZBPLZ = 2
ZBOrt = 3
ZBStrasse = 4
Bereich = 5
NameJM = 6
VornameJM = 7
OrtJM = 8
Massnahme = 9
Hilfe_aus_BH = 10
Kassenzeichen = 11
Buchungsdatum = 12
Betrag = 13
Verwendungszweck_1 = 14
Verwendungszweck_2 = 15
Bezeichnung = 16
Ergebnis = 17
Verkettung_TPP = 18
Verkettung_TK = 19
Verkettung_Wohnort = 20
fuer_Monat = 21
Buchungstag = 22
Buchungsmonat = 23
Buchungsjahr = 24 #bzw. +1 wenn Jahreswechsel!
Buchungsmonat_Text = 25
Zwischenberechnung_Zuflussmonat = 26
Zuflussmonat = 27
Cluster_erh_FB = 28
Cluster_Vertretung = 29
Cluster_AU = 30
Ort = 31
SVERWEIS_Pruefung_TPP = 32
TPP = 33
RUE_TPP = 34
Hilfsspalte = 35

L_DATE_INX = [Buchungsdatum]
L_FLOAT_INX = [Betrag, Ergebnis]

L_SECOND_TABLE_KEYS = ["außergewöhnlich", "Vertretung", "erhöhter Förderbedarf"]
S_FIRST_TABLE_DICT = "normal"
S_SECOND_TABLE_DICT = "special"

L_MONTH_NAME = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]

S_SHEET_EK_1HJ = "EK-Berechnung (1. HJ)"
S_SHEET_EK_2HJ = "EK-Berechnung (2. HJ)"
S_SHEET_ZU_1HJ = "Zuschuss (1. HJ)"
S_SHEET_ZU_2HJ = "Zuschuss (2. HJ)"

I_TABLE_OFFSET_TOP = 0
I_TABLE_OFFSET_BOTTOM = 20

COLOR_LIGHTGREEN = "CCFFCC"
COLOR_LIGHTBLUE = "B8CCE4"
COLOR_GREY = "A6A6A6"
COLOR_LIGHTORANGE = "FCD5B4"

S_EUR_FORMAT = '#,##0.00 €'
S_PERCENT_FORMAT = '0.00%'

S_PERCENT_CONDITION = "<2019"
S_PERCENT_1 = "68.3%"
S_PERCENT_2 = "73.2%"

F_MIN_REFUND = 470
F_MAX_REFUND = 1097.67
F_MIN_PAY = 186.96
F_MIN_PAY_WITH_KTG = 193.56
F_MIN_AV_REFUND = 43
F_AV_LIMIT = 450

F_KV = 14.6/100
F_PV = 3.05/100
F_PV_KL = 3.4/100
F_ZU = 0/100
F_AV = 18.6/100
F_DYN = 1 # dynamic factor

I_DEFAULT_DAY = 25
# Billing day for next month
D_MONTH_BILLING_DATE = {
   1:  I_DEFAULT_DAY, # January
   2:  I_DEFAULT_DAY, # February
   3:  I_DEFAULT_DAY, # March
   4:  I_DEFAULT_DAY, # April
   5:  I_DEFAULT_DAY, # May
   6:  23, # June
   7:  I_DEFAULT_DAY, # July
   8:  I_DEFAULT_DAY, # August
   9:  I_DEFAULT_DAY, # September
   10: I_DEFAULT_DAY, # October
   11: I_DEFAULT_DAY, # November
   12: 22  # December
}

I_HALF_MONTH_YEAR = 6
I_MONTH_COLUMN_OFFSET = 4
I_MAX_CHILDS = 12
I_CHILD_OFFSET = 5

THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

DOUBLE_UNDERLINED_BORDER = Border(bottom=Side(style='double'))

DOUBLE_BORDER = Border(left=Side(style='double'),
                     right=Side(style='double'),
                     top=Side(style='double'),
                     bottom=Side(style='double'))

F_SCALE_FACTOR = 22.43/21.71

# App Data
S_WJHSV_APPLICATION_NAME = "WJH-SV"
S_WJHSV_DESCRIPTION = "Wirtschaftliche Jugendhilfe - Sozialversicherung"
I_VERSION_NUM_1 = 1
I_VERSION_NUM_2 = 0
I_VERSION_NUM_3 = 1
S_VERSION = str(I_VERSION_NUM_1) + '.' + str(I_VERSION_NUM_2) + '.' + str(I_VERSION_NUM_3)
S_COPYRIGHT = "Copyright © 2022 Timo Unger"
S_LICENSE = "GNU General Public License"
S_HOME = "https://timounger.github.io/WJH-SV"
S_APP_ID = S_WJHSV_APPLICATION_NAME + '.' + S_VERSION
S_ICON_RESOURCE_PATH = 'Resources/wjh_sv.png'

class SubsidyCalculator():
    """!
    @brief Class SubsidyCalculator: calculate social insurance subsidy for supervisor
    """
    def __init__(self):
        self.run_calculation()

    def run_calculation(self):
        """!
        @brief Run calculation of WJH-SV
        """
        args = self.args_parser()
        if args is not None or B_DEBUG:
            if not B_DEBUG:
                i_year = args.year
                s_file = args.file
                s_sheet = args.sheet
            else:
                i_year = I_YEAR
                s_file = S_FILE
                s_sheet = S_SHEET
            if os.path.isfile(s_file):
                now = datetime.now()
                l_table = self.get_table_data(s_file, s_sheet)
                l_user = self.get_user(l_table)
                if l_user:
                    s_output_folder = S_OUTPUT_FOLDER
                    if not B_DEBUG:
                        s_output_folder += "/" + now.strftime("SV_Berechnung" + "_%Y-%m-%d_%Hh%Mm%Ss")
                    if not os.path.exists(s_output_folder):
                        os.makedirs(s_output_folder)
                for l_user_data in l_user:
                    #if not B_DEBUG or l_user_data[0] == "Unger" and l_user_data[1] == "Tamara":
                    self.create_report_file(s_output_folder, l_table, i_year, l_user_data[0], l_user_data[1])
                print("Finished!")
            else:
                sys.exit(f"File not exist: {s_file}")

    def args_parser(self):
        """!
        @brief Get script arguments.
        @return all present script arguments; None for no arguments
        """
        if len( sys.argv ) > 1:
            my_parser = argparse.ArgumentParser()
            required_args = my_parser.add_argument_group('required named arguments')
            required_args.add_argument('--year', type=int, help=f'Year of calculation (e.g. {I_YEAR})')
            required_args.add_argument('--file', type=str, help=f'file with input data (e.g. {S_FILE})')
            required_args.add_argument('--sheet', type=str, help=f'sheet name of file with input data (e.g. {S_SHEET})')
            args = my_parser.parse_args()
        else:
            print("Arguments missing. Use --help for more details.")
            args = None
        return args

    def get_table_data(self, s_file, s_sheet):
        """!
        @brief get all data from input file.
        @param s_file : input file name
        @param s_sheet : input sheet name
        """
        print(f"Read file: {s_file} ...")
        workbook = openpyxl.load_workbook(s_file)
        table_sheet = workbook[s_sheet]
        l_table = []
        for i_row, row in enumerate(table_sheet.iter_rows()):
            l_data_row = []
            if row[0].value is not None:
                if i_row == I_HEADER_ROW_INX: # ignore title of columns
                    if B_DEBUG:
                        l_header = []
                        for cell in row:
                            l_header.append(str(cell.value))
                else:
                    for i_cell, cell in enumerate(row):
                        if cell.value is None:
                            l_data_row.append("")
                        elif i_cell in L_FLOAT_INX:
                            l_data_row.append(float(cell.value))
                        elif i_cell in L_DATE_INX:
                            datetime_object = datetime.strptime(str(cell.value), '%Y-%m-%d %H:%M:%S')
                            l_data_row.append(datetime_object)
                        else:
                            l_data_row.append(str(cell.value))
                    l_table.append(l_data_row)
            else:
                if i_row != I_EMPTY_ROW_INX: # ignore first empty row
                    break
        return l_table

    def get_user(self, l_table):
        """!
        @brief Get list of user in input file
        @param l_table : list with all data from input file
        @return list with user in following format [[Last Name, First Name], [], ... ]
        """
        l_user = []
        for entry in l_table:
            l_user_name = [entry[ZBNachname], entry[ZBVorname]]
            if l_user_name not in l_user:
                l_user.append(l_user_name)
        return l_user

    def create_report_file(self, s_output_folder, l_table, i_year, s_name, s_first_name):
        """!
        @brief Create report file of an user
        @param s_output_folder : output folder
        @param l_table : data from input file
        @param i_year : required year
        @param s_name : last name of user
        @param s_first_name : first name of user
        """
        d_data = self.calculate_ek(l_table, i_year, s_name, s_first_name)
        workbook = Workbook()
        # sheet with calculation 1.HJ
        ws_ek_1 = workbook.active
        ws_ek_1.title = S_SHEET_EK_1HJ
        self.create_calculation_sheet(ws_ek_1, d_data, i_year, s_name, s_first_name, True)
        # sheet with calculation 2.HJ
        ws_ek_2 = workbook.create_sheet(S_SHEET_EK_2HJ)
        self.create_calculation_sheet(ws_ek_2, d_data, i_year, s_name, s_first_name, False)
        # sheet with subsidy 1.HJ
        ws_zu_1 = workbook.create_sheet(S_SHEET_ZU_1HJ)
        self.create_subsidy_sheet(ws_zu_1, i_year, True)
        #workbook.active = workbook[S_SHEET_ZU_1HJ]
        # sheet with subsidy 2.HJ
        ws_zu_2 = workbook.create_sheet(S_SHEET_ZU_2HJ)
        self.create_subsidy_sheet(ws_zu_2, i_year, False)
        # save file
        s_file_name = f"{s_output_folder}/SV_Berechnung_{i_year}_{s_name}_{s_first_name}.xlsx"
        workbook.save(filename = s_file_name)
        print(f"Created: {s_file_name}")

    def calculate_ek(self, l_table, i_year, s_name, s_first_name):
        """!
        @brief Create report file of an user
        @param l_table : data from input file
        @param i_year : required year
        @param s_name : last name of user
        @param s_first_name : first name of user
        @return dictionary with relevant data of user
        """
        d_data = {S_FIRST_TABLE_DICT: {}, S_SECOND_TABLE_DICT: {}}
        for entry in l_table:
            i_entry_year = entry[Buchungsdatum].year
            i_entry_month = entry[Buchungsdatum].month
            i_entry_day = entry[Buchungsdatum].day
            if (entry[ZBNachname] == s_name) and (entry[ZBVorname] == s_first_name):
                i_real_month = i_entry_month
                i_real_year = i_entry_year
                if (i_entry_day >= D_MONTH_BILLING_DATE[i_entry_month]):
                    i_real_month += 1
                    if i_real_month == 13:
                        i_real_month = 1
                        i_real_year += 1
                if i_real_year == i_year:
                    s_child = f"{entry[NameJM]}, {entry[VornameJM]}, {entry[OrtJM]}"
                    s_table_part = S_FIRST_TABLE_DICT
                    for s_special in L_SECOND_TABLE_KEYS:
                        if s_special.lower() in entry[Bezeichnung].lower():
                            s_table_part = S_SECOND_TABLE_DICT
                            break
                    if s_child not in d_data[s_table_part]:
                        d_data[s_table_part][s_child] = {}
                    if i_real_month not in d_data[s_table_part][s_child]:
                        d_data[s_table_part][s_child][i_real_month] = float(0)
                    d_data[s_table_part][s_child][i_real_month] += entry[Betrag]
        return d_data

    def create_calculation_sheet(self, ws, d_data, i_year, s_name, s_first_name, b_first_half_year = True):
        """!
        @brief Create calculation sheet
        @param ws : actual worksheet
        @param d_data : dictionary with relevant user of data
        @param i_year : required year
        @param s_name : last name of user
        @param s_first_name : first name of user
        @param b_first_half_year : [True] calculation of first half year; [False] second half year
        """
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.scale = 70 # in percent TODO try auto scale
        # set row high
        l_row_high = [30, 25, 5] + ([13] * 6) + [16]
        for i, f_row_width in enumerate(l_row_high, start=1):
            ws.column_dimensions[get_column_letter(i)].width = f_row_width * F_SCALE_FACTOR
        # set data
        self.create_calculation_header(ws, i_year, s_name, s_first_name)
        self.create_calculation_table(ws, d_data, b_first_half_year, True)
        self.create_calculation_sum(ws, True)
        self.create_calculation_table(ws, d_data, b_first_half_year, False)
        self.create_calculation_sum(ws, False)

    def create_calculation_header(self, ws, i_year, s_name, s_first_name):
        """!
        @brief Create header in calculation sheet
        @param ws : actual worksheet
        @param i_year : required year
        @param s_name : last name of user
        @param s_first_name : first name of user
        """
        self.set_cell(ws, 'A1', "Einkommensberechnung", b_bold=True)
        self.set_cell(ws, 'C1', f"{s_name}, {s_first_name}", fill_color=COLOR_LIGHTGREEN)
        self.set_cell(ws, 'D1', fill_color=COLOR_LIGHTGREEN)
        self.set_cell(ws, 'F1', "Jahr", b_bold=True)
        self.set_cell(ws, 'G1', i_year, b_bold=True, align='left', fill_color=COLOR_LIGHTGREEN)

    def create_calculation_table(self, ws, d_data, b_first_half_year = True, b_normal_table = True):
        """!
        @brief Create table in calculation sheet
        @param ws : actual worksheet
        @param d_data : dictionary with relevant user of data
        @param b_first_half_year : [True] first half year; [False] second half year
        @param b_normal_table : [True] normal table; [False] special table
        """
        if b_first_half_year:
            i_year_offset = 0
            l_valid_months = list(range(1, I_HALF_MONTH_YEAR+1))
        else:
            i_year_offset = I_HALF_MONTH_YEAR
            l_valid_months = list(range(I_HALF_MONTH_YEAR+1, (I_HALF_MONTH_YEAR*2)+1))
        if b_normal_table:
            i_offset = I_TABLE_OFFSET_TOP
            s_table_part = S_FIRST_TABLE_DICT
        else:
            i_offset = I_TABLE_OFFSET_BOTTOM
            s_table_part = S_SECOND_TABLE_DICT
        for i_month in range(I_HALF_MONTH_YEAR):
            self.set_cell(ws, get_column_letter(i_month + I_MONTH_COLUMN_OFFSET) + str(i_offset+3), L_MONTH_NAME[i_year_offset+i_month], b_bold=True, align='center')
        self.set_cell(ws, 'J' + str(i_offset+3), "Gesamt", b_bold=True, align='center')
        self.set_cell(ws, 'A' + str(i_offset+4), "Betreute Kinder", b_bold=True)
        self.set_cell(ws, 'B' + str(i_offset+4), "Wohnort", b_bold=True)
        d_filtered_user_data = {}
        for i, s_child in enumerate(d_data[s_table_part]): # filter to sort out user only for other half year
            if any(key in d_data[s_table_part][s_child] for key in l_valid_months):
                d_filtered_user_data[s_child] = d_data[s_table_part][s_child]
        i_row = 0
        for i, s_child in enumerate(d_filtered_user_data):
            i_row = i
            i_index = s_child.rfind(',')
            self.set_cell(ws, 'A' + str(i_offset+I_CHILD_OFFSET+i), s_child[:i_index], fill_color=COLOR_LIGHTGREEN)
            self.set_cell(ws, 'B' + str(i_offset+I_CHILD_OFFSET+i), s_child[i_index + 2:], fill_color=COLOR_LIGHTBLUE)
            self.set_cell(ws, 'C' + str(i_offset+I_CHILD_OFFSET+i), fill_color=COLOR_GREY)
            for i_month in range(I_HALF_MONTH_YEAR):
                i_real_month = i_year_offset + i_month + 1
                if i_real_month in d_filtered_user_data[s_child]:
                    f_sum = d_filtered_user_data[s_child][i_real_month]
                else:
                    f_sum = 0
                s_cell = get_column_letter(i_month + I_MONTH_COLUMN_OFFSET) + str(i_offset+I_CHILD_OFFSET+i)
                self.set_cell(ws, s_cell, f_sum, fill_color=COLOR_LIGHTORANGE)
                ws[s_cell].number_format = S_EUR_FORMAT
        if d_filtered_user_data:
            s_start_row = i_row + 1
        else:
            s_start_row = i_row
        for i in range(s_start_row, I_MAX_CHILDS):
            self.set_cell(ws, 'A' + str(i_offset+I_CHILD_OFFSET+i), fill_color=COLOR_LIGHTGREEN)
            self.set_cell(ws, 'B' + str(i_offset+I_CHILD_OFFSET+i), fill_color=COLOR_LIGHTBLUE)
            self.set_cell(ws, 'C' + str(i_offset+I_CHILD_OFFSET+i), fill_color=COLOR_GREY)
            for i_month in range(I_HALF_MONTH_YEAR):
                s_cell = get_column_letter(i_month + I_MONTH_COLUMN_OFFSET) + str(i_offset+I_CHILD_OFFSET+i)
                self.set_cell(ws, s_cell, 0.00, fill_color=COLOR_LIGHTORANGE, s_format=S_EUR_FORMAT)
        for i in range(I_MAX_CHILDS):
            s_line = str(i_offset+I_CHILD_OFFSET+i)
            s_cell = 'J' + s_line
            self.set_cell(ws, s_cell, f'=SUM(D{s_line}:I{s_line})', fill_color=COLOR_LIGHTORANGE, s_format=S_EUR_FORMAT)
        # set border
        for i_colum in range(10):
            for i_row in range(I_MAX_CHILDS + 2):
                s_cell = get_column_letter(i_colum + 1) + str(i_offset+3+i_row)
                ws[s_cell].border = THIN_BORDER

    def create_calculation_sum(self, ws, b_normal_table = True):
        """!
        @brief Create sum in calculation sheet
        @param ws : actual worksheet
        @param b_normal_table : [True] normal table; [False] special table
        """
        if b_normal_table:
            i_offset = I_TABLE_OFFSET_TOP
            s_text_1 = 'Gesamtsumme Pflegegeld 1. Halbjahr ='
            s_text_2 = 'durchschnittliches reguläres Pflegegeld mtl. ='
            s_text_3 = 'steuerpflichtiges Einkommen ='
            s_formula = f'=IF(G1{S_PERCENT_CONDITION}, J19*{S_PERCENT_1}, J19*{S_PERCENT_2})'
            i_merge_start = 6
            s_column = 'F'
            border = DOUBLE_UNDERLINED_BORDER
        else:
            i_offset = I_TABLE_OFFSET_BOTTOM
            s_text_1 = 'Gesamtsumme erhöhter Förderbedarf/Vertretung/außergewöhnliche Betreuungszeiten 1. Halbjahr ='
            s_text_2 = 'durchschnittliches zusätzliches Pflegegeld mtl. ='
            s_text_3 = 'zu berücksichtigendes Einkommen = '
            s_formula = '=J39+J21'
            i_merge_start = 1
            s_column = 'A'
            border = DOUBLE_BORDER
        self.set_cell(ws, s_column + str(i_offset+18), s_text_1, b_bold=True, align='right')
        ws.merge_cells(start_row=i_offset+18, start_column=i_merge_start, end_row=i_offset+18, end_column=9)
        self.set_cell(ws, 'J' + str(i_offset+18), f'=SUM(J{str(i_offset+I_CHILD_OFFSET)}:J{str(i_offset+I_CHILD_OFFSET + I_MAX_CHILDS - 1)})', b_bold=True, fill_color=COLOR_LIGHTBLUE, s_format=S_EUR_FORMAT)
        if b_normal_table:
            s_month_count_formula = "=SUM(IF(SUM(D5:D16,D25:D36)<>0,1,0)+IF(SUM(E5:E16,E25:E36)<>0,1,0)+IF(SUM(F5:F16,F25:F36)<>0,1,0)+IF(SUM(G5:G16,G25:G36)<>0,1,0)+IF(SUM(H5:H16,H25:H36)<>0,1,0)+IF(SUM(I5:I16,I25:I36)<>0,1,0))"
            self.set_cell(ws, 'B' + str(i_offset+19), "teilen durch Monate", b_bold=True, align='right')
            self.set_cell(ws, 'C19', s_month_count_formula, b_bold=True, fill_color=COLOR_LIGHTGREEN)
            self.set_cell(ws, 'E19', '←')
        self.set_cell(ws, s_column + str(i_offset+19), s_text_2, b_bold=True, align='right')
        ws.merge_cells(start_row=i_offset+19, start_column=i_merge_start, end_row=i_offset+19, end_column=9)
        self.set_cell(ws, 'J' + str(i_offset+19), f'=IFERROR(J{str(i_offset+18)}/C19, 0)', b_bold=True, fill_color=COLOR_LIGHTBLUE, s_format=S_EUR_FORMAT, s_border=DOUBLE_UNDERLINED_BORDER)
        self.set_cell(ws, 'F' + str(i_offset+21), s_text_3, b_bold=True, align='right')
        ws.merge_cells(start_row=i_offset+21, start_column=6, end_row=i_offset+21, end_column=9)
        self.set_cell(ws, 'J' + str(i_offset+21), s_formula, b_bold=True, fill_color=COLOR_LIGHTBLUE, s_format=S_EUR_FORMAT, s_border=border)

    def create_subsidy_sheet(self, ws, i_year, b_first_half_year = True):
        """!
        @brief Create subsidy sheet
        @param ws : actual worksheet
        @param i_year : required year
        @param b_first_half_year : [True] first half year; [False] second half year
        """
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.scale = 88 # in percent TODO try auto scale
        ws.sheet_properties.outlinePr.summaryBelow = False
        # set row high
        l_row_high = [37.57, 11.14, 16, 7, 11.71]
        for i, f_row_width in enumerate(l_row_high, start=1):
            ws.column_dimensions[get_column_letter(i)].width = f_row_width * F_SCALE_FACTOR
        # set data
        if b_first_half_year:
            s_ek_name = S_SHEET_EK_1HJ
            s_part_year = "1"
        else:
            s_ek_name = S_SHEET_EK_2HJ
            s_part_year = "2"
        self.set_cell(ws, 'A1', f"Berechnung für Kranken- und Pflegeversicherung (KV, PV) - {s_part_year}. Halbjahr", b_bold=True, b_underline=True, fill_color=COLOR_LIGHTBLUE)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        i_table_offset = 11
        i_start_offset = 2
        i_offset = 0
        self.create_subsidy_kp_part(ws, i_year, s_ek_name, i_offset)
        i_offset += i_table_offset
        self.create_subsidy_kp_part(ws, i_year, s_ek_name, i_offset)
        i_offset += i_table_offset
        self.create_subsidy_kp_part(ws, i_year, s_ek_name, i_offset)
        ws.row_dimensions.group(3, hidden=True)
        ws.row_dimensions.group(i_start_offset+i_table_offset, i_start_offset+(3*i_table_offset)-1, hidden=True, outline_level=1)
        ws.row_dimensions.group(i_start_offset+(2*i_table_offset), i_start_offset+(3*i_table_offset)-1, hidden=True, outline_level=2)

        self.set_cell(ws, f'A{14+i_offset}', "Insgesamt werden erstattet:", b_bold=True, align='right', i_font_size = 10)
        i_total_row = 10
        self.set_cell(ws, f'B{14+i_offset}', f"=((B{i_total_row}*D{i_total_row})+(B{i_total_row+i_table_offset}*D{i_total_row+i_table_offset})+(B{i_total_row+(2*i_table_offset)}*D{i_total_row+(2*i_table_offset)}))", b_bold=True, fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=DOUBLE_UNDERLINED_BORDER)
        self.set_cell(ws, f'A{16+i_offset}', "Hinweise:", b_bold=True, b_underline=True, i_font_size=10)
        l_text = [f"> keine Versicherungspflicht bei einem zu versteuerndem EK unter {F_MIN_REFUND} €.",
                  f"> immer Mindestbeitrag zwischen {F_MIN_REFUND} € und {F_MAX_REFUND} € (zu versteuerndem EK).",
                  f"> Mindestbeitrag KV ohne Krankentagegeldversicherung: 14,0% = mindestens {F_MIN_PAY} € (inkl. PV)",
                  f"> Mindestbeitrag KV mit Krankentagegeldversicherung: {F_KV:.1%} = mindestens {F_MIN_PAY_WITH_KTG} € (inkl. PV)",
                  f"> pflichtversichert über einem EK von {F_MAX_REFUND} €  (KV = {F_KV:.1%} und PV = {F_PV:.2%} bzw. {F_PV_KL:.2%} bei kinderlos).",
                  "(1) liegt der KK ein Steuerbescheid bspw. für Jahr 2018 vor, ermittelt sie auf dieser Grundlage die Beiträge.",
                  "Das Pflegegeld aus dem Jahr 2018 ist dann für die Berechnung der Erstattung heranzuziehen."]
        for i, s_text in enumerate(l_text):
            self.set_cell(ws, f'A{17+i+i_offset}', s_text, i_font_size=10)

        self.set_cell(ws, f'A{25+i_offset}', f"Berechnung für Altersvorsorge (AV) - {s_part_year}. Halbjahr", b_bold=True, b_underline=True, fill_color=COLOR_LIGHTBLUE)
        ws.merge_cells(start_row=(25+i_offset), start_column=1, end_row=(25+i_offset), end_column=5)
        i_table_offset = 10
        i_start_offset = 58
        self.create_subsidy_av_part(ws, i_year, s_ek_name, i_offset)
        i_offset += i_table_offset
        self.create_subsidy_av_part(ws, i_year, s_ek_name, i_offset)
        i_offset += i_table_offset
        self.create_subsidy_av_part(ws, i_year, s_ek_name, i_offset)
        ws.row_dimensions.group(49, hidden=True)
        ws.row_dimensions.group(i_start_offset, i_start_offset+(2*i_table_offset)-1, hidden=True, outline_level=1)
        ws.row_dimensions.group(i_start_offset+i_table_offset, i_start_offset+(2*i_table_offset)-1, hidden=True, outline_level=2)

        self.set_cell(ws, f'A{37+i_offset}', "Insgesamt werden erstattet:", b_bold=True, align='right', i_font_size = 10)
        i_total_row = 57
        self.set_cell(ws, f'B{37+i_offset}', f"=((B{i_total_row}*D{i_total_row})+(B{i_total_row+i_table_offset}*D{i_total_row+i_table_offset})+(B{i_total_row+(2*i_table_offset)}*D{i_total_row+(2*i_table_offset)}))", b_bold=True, fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=DOUBLE_UNDERLINED_BORDER)
        self.set_cell(ws, f'A{39+i_offset}', "Hinweise:", b_bold=True, b_underline=True, i_font_size=10)
        l_text = [f"> nicht Versicherungspflichtig bei einem zu versteuerndem EK unter {F_AV_LIMIT} €.",
                  f"> kann aber auf Nachweis mit maximal {F_MIN_AV_REFUND} € gefördert werden.",
                  f"> versicherungspflichtig bei einem EK über {F_AV_LIMIT} €.",
                  "(1) liegt der RV ein Steuerbescheid bspw. für Jahr 2018 vor, ermittelt sie auf seiner Grundlage die Beiträge.",
                  "Das Pflegegeld des gleichen Jahres 2018 ist dann für die Berechnung der Erstattung heranzuziehen."]
        for i, s_text in enumerate(l_text):
            self.set_cell(ws, f'A{40+i+i_offset}', s_text, i_font_size=10)

        self.set_cell(ws, f'A{46+i_offset}', "Unfallversicherung (UV)", b_bold=True, b_underline=True, fill_color=COLOR_LIGHTBLUE)
        ws.merge_cells(start_row=46+i_offset, start_column=1, end_row=46+i_offset, end_column=5)
        self.set_cell(ws, f'B{48+i_offset}', "Jahr:", align='right', i_font_size = 10)
        self.set_cell(ws, f'C{48+i_offset}', fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10) # TODO ausfüllen
        self.set_cell(ws, f'B{50+i_offset}', "Betrag:", align='right', i_font_size = 10)
        self.set_cell(ws, f'C{50+i_offset}', fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10) # TODO ausfüllen

        self.set_cell(ws, f'A{53+i_offset}', f"Erstattung {s_part_year}. Halbjahr:", b_bold=True, b_italic=True, b_underline=True, i_font_size = 14, fill_color=COLOR_GREY)
        self.set_cell(ws, f'B{53+i_offset}', fill_color=COLOR_GREY, b_bold=True, b_italic=True, i_font_size = 14)
        self.set_cell(ws, f'C{53+i_offset}', "=B36+B79+C92", b_bold=True, b_italic=True, b_underline=True, fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 14)
        self.set_cell(ws, f'D{53+i_offset}', b_bold=True, b_italic=True, fill_color=COLOR_GREY, i_font_size = 14)
        self.set_cell(ws, f'E{53+i_offset}', b_bold=True, b_italic=True, fill_color=COLOR_GREY, i_font_size = 14)
        
    def create_subsidy_kp_part(self, ws, i_year, s_ek_name, i_offset):
        """!
        @brief Create subsidy sheet
        @param ws : actual worksheet
        @param i_year : required year
        @param s_ek_name : ek sheet name
        @param i_offset : cell offset
        """
        self.set_cell(ws, f'A{3+i_offset}', f"ab 01.01.{i_year}", i_font_size = 10)
        self.set_cell(ws, f'A{4+i_offset}', "maßgebl. Jahr der Beitragsbemessung(1):", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{4+i_offset}', i_year, fill_color=COLOR_GREY, i_font_size = 10, s_border=THIN_BORDER) # TODO individuell ausfüllen wenn Tagespflege Liste besteht (anhand letztem Steuerbescheid)
        self.set_cell(ws, f'A{5+i_offset}', "durchschn. mtl. Pflegegeld(1):", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{5+i_offset}', f"='{s_ek_name}'!J19", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'A{6+i_offset}', "steuerpflichtiges Einkommen aus TP(1):", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{6+i_offset}', f"=IF(B{4+i_offset}{S_PERCENT_CONDITION}, B{5+i_offset}*{S_PERCENT_1}, B{5+i_offset}*{S_PERCENT_2})", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'A{7+i_offset}', "zu berücksichtigendes Einkommen aus TP(1):", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{7+i_offset}', f"='{s_ek_name}'!J41", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)

        self.set_cell(ws, f'C{4+i_offset}', "KV =", align='right', i_font_size = 10)
        self.set_cell(ws, f'D{4+i_offset}', F_KV, fill_color=COLOR_GREY, s_format=S_PERCENT_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'E{4+i_offset}', f"=(B{7+i_offset}*D{4+i_offset})", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'C{5+i_offset}', "PV =", align='right', i_font_size = 10)
        self.set_cell(ws, f'D{5+i_offset}', F_PV, fill_color=COLOR_GREY, s_format=S_PERCENT_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'E{5+i_offset}', f"=(B{7+i_offset}*D{5+i_offset})", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'C{6+i_offset}', "Zusatzbeitrag =", align='right', i_font_size = 10)
        self.set_cell(ws, f'D{6+i_offset}', F_ZU, fill_color=COLOR_GREY, s_format=S_PERCENT_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'E{6+i_offset}', f"=(B{7+i_offset}*D{6+i_offset})", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'C{7+i_offset}', "Gesamt =", align='right', i_font_size = 10)
        self.set_cell(ws, f'D{7+i_offset}', f"=SUM(E{4+i_offset}:E{6+i_offset})", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        ws.merge_cells(start_row=(7+i_offset), start_column=4, end_row=(7+i_offset), end_column=5)

        self.set_cell(ws, f'A{9+i_offset}', "aus Beitragsbescheid Krankenkasse:", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{9+i_offset}', fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER) # TODO ausfüllen
        self.set_cell(ws, f'A{10+i_offset}', "Erstattungsbetrag (mtl.) durch LRA:", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{10+i_offset}', f"=IF(AND(B{6+i_offset}>={F_MIN_REFUND},B{6+i_offset}<={F_MAX_REFUND}),MIN({F_MIN_PAY}/2,B{9+i_offset}/2),IF(B{9+i_offset}>D{7+i_offset},D{7+i_offset}/2,B{9+i_offset}/2))", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'C{10+i_offset}', "Anzahl Monate:", align='right', i_font_size = 10)
        self.set_cell(ws, f'D{10+i_offset}', f"='{s_ek_name}'!C19", fill_color=COLOR_GREY, i_font_size = 10)
        self.set_cell(ws, f'A{11+i_offset}', "davon für KV:", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{11+i_offset}', f"=B{10+i_offset}/(D{4+i_offset}+D{5+i_offset}+D{6+i_offset})*(D{4+i_offset}+D{6+i_offset})", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'A{12+i_offset}', "davon für PV:", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{12+i_offset}', f"=B{10+i_offset}/(D{4+i_offset}+D{6+i_offset}+D{5+i_offset})*D{5+i_offset}", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)

    def create_subsidy_av_part(self, ws, i_year, s_ek_name, i_offset):
        """!
        @brief Create subsidy sheet
        @param ws : actual worksheet
        @param i_year : required year
        @param s_ek_name : ek sheet name
        @param i_offset : cell offset
        """
        self.set_cell(ws, f'A{27+i_offset}', f"ab 01.01.{i_year}", i_font_size = 10)
        self.set_cell(ws, f'A{28+i_offset}', "maßgebl. Jahr der Beitragsbemessung(1):", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{28+i_offset}', i_year, fill_color=COLOR_GREY, i_font_size = 10, s_border=THIN_BORDER) # TODO individuell ausfüllen wenn Tagespflege Liste besteht (anhand letztem Steuerbescheid)
        self.set_cell(ws, f'A{29+i_offset}', "durchschn. mtl. Pflegegeld(1):", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{29+i_offset}', f"='{s_ek_name}'!J19", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'A{30+i_offset}', "dynamisiert:", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{30+i_offset}', f"=B{29+i_offset}*D{30+i_offset}", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'A{31+i_offset}', "steuerpflichtiges Einkommen aus TP(1):", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{31+i_offset}', f"=IF(B{28+i_offset}{S_PERCENT_CONDITION}, B{30+i_offset}*{S_PERCENT_1}, B{30+i_offset}*{S_PERCENT_2})", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)

        self.set_cell(ws, f'A{32+i_offset}', "zu berücksichtigendes Einkommen(1):", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{32+i_offset}', f"=(('{s_ek_name}'!J39*D{30+i_offset})+B{30+i_offset})", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)

        self.set_cell(ws, f'C{29+i_offset}', "AV =", align='right', i_font_size = 10)
        self.set_cell(ws, f'D{29+i_offset}', F_AV, fill_color=COLOR_GREY, s_format=S_PERCENT_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'E{29+i_offset}', f"=(B{32+i_offset}*D{29+i_offset})", fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER)
        self.set_cell(ws, f'C{30+i_offset}', "dyn. Faktor =", align='right', i_font_size = 10)
        self.set_cell(ws, f'D{30+i_offset}', F_DYN, fill_color=COLOR_GREY, i_font_size = 10, s_border=THIN_BORDER)
        ws.merge_cells(start_row=30+i_offset, start_column=4, end_row=30+i_offset, end_column=5)

        self.set_cell(ws, f'A{34+i_offset}', "aus Beitragsbescheid Rentenversicherung:", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{34+i_offset}', fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER) # TODO ausfüllen
        self.set_cell(ws, f'A{35+i_offset}', "Erstattungsbetrag (mtl.) durch LRA:", align='right', i_font_size = 10)
        self.set_cell(ws, f'B{35+i_offset}', f"=MIN(MAX(IF(B{34+i_offset}>E{29+i_offset},E{29+i_offset}/2,B{34+i_offset}/2),{F_MIN_AV_REFUND}),B{34+i_offset}/2)",fill_color=COLOR_GREY, s_format=S_EUR_FORMAT, i_font_size = 10, s_border=THIN_BORDER) # TODO ausfüllen
        self.set_cell(ws, f'C{35+i_offset}', "Anzahl Monate:", align='right', i_font_size = 10)
        self.set_cell(ws, f'D{35+i_offset}', f"='{s_ek_name}'!C19", fill_color=COLOR_GREY, i_font_size = 10)

    def set_cell(self, ws, s_cell, value = None, b_bold = False, b_italic=False, b_underline = False, i_font_size = 12, s_font = 'Arial', fill_color = None, align = None, s_format = None, s_border = None):
        """!
        @brief Set cell data
        @param ws : actual worksheet
        @param s_cell : cell to set data
        @param value : value to set in cell; None: set no data to cell
        @param b_bold : status if cell content should be bold
        @param b_italic : status if cell content should be italic
        @param b_underline : status if cell content should be underlined
        @param i_font_size : font size
        @param s_font : font art
        @param fill_color : backgourn fill color of cell
        @param align : text align of cell
        @param s_format : format of cell
        @param s_border : boarder of cell
        """
        if value is not None:
            ws[s_cell].value = value
        if b_underline:
            s_underline = 'single'
        else:
            s_underline = 'none'
        ws[s_cell].font = Font(name = s_font, size = str(i_font_size), bold=b_bold, italic=b_italic, underline=s_underline)
        if fill_color is not None:
            ws[s_cell].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type = "solid")
        if align is not None:
            ws[s_cell].alignment = Alignment(horizontal=align)
        if s_format is not None:
            ws[s_cell].number_format = s_format
        if s_border is not None:
            ws[s_cell].border = s_border

if __name__ == "__main__":
    print(f"{S_WJHSV_APPLICATION_NAME} - Version: {S_VERSION}")
    print(S_WJHSV_DESCRIPTION)
    #print("")
    print(S_COPYRIGHT)
    #print(S_LICENSE)
    #print(f"Home: {S_HOME}")
    print("")

    SubsidyCalculator()

    sys.exit()
